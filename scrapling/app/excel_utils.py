"""Excel utilities for generating formatted product export files."""
import os
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Exact column order matching the user's CSV template
CSV_TEMPLATE_COLUMNS = [
    'Category',
    'Product Name',
    'Brand',
    'Manufacturer',
    'Sale Price',
    'Discount Base Price',
    'Stock',
    'Lead Time',
    'Detailed Description',
    'Main Image',
    'Search Keywords',
    'Quantity',
    'Volume',
    'Weight',
    'Adult Only',
    'Taxable',
    'Parallel Import',
    'Overseas Purchase',
    'SKU',
    'Model Number',
    'Barcode',
    'Additional Image 1',
    'Additional Image 2',
    'Additional Image 3',
    'Additional Image 4',
    'Additional Image 5',
    'Product URL',
]

# Column width configuration
COLUMN_WIDTHS = {
    'Category': 15,
    'Product Name': 45,
    'Brand': 20,
    'Manufacturer': 20,
    'Sale Price': 12,
    'Discount Base Price': 18,
    'Stock': 10,
    'Lead Time': 12,
    'Detailed Description': 60,
    'Main Image': 50,
    'Search Keywords': 40,
    'Quantity': 10,
    'Volume': 12,
    'Weight': 12,
    'Adult Only': 10,
    'Taxable': 10,
    'Parallel Import': 12,
    'Overseas Purchase': 15,
    'SKU': 18,
    'Model Number': 18,
    'Barcode': 18,
    'Additional Image 1': 50,
    'Additional Image 2': 50,
    'Additional Image 3': 50,
    'Additional Image 4': 50,
    'Additional Image 5': 50,
    'Product URL': 45,
}

# URL columns that should be styled as links
URL_COLUMNS = {'Main Image', 'Additional Image 1', 'Additional Image 2', 'Additional Image 3', 'Additional Image 4', 'Additional Image 5', 'Product URL'}

def build_excel(products, keyword, base_url, outputs_dir, partial=False):
    """Build formatted Excel workbook from scraped product data.
    
    Args:
        products: List of product dicts to export
        keyword: Search keyword used for scraping
        base_url: Base URL of the e-commerce site
        outputs_dir: Directory to save the Excel file
        partial: If True, marks the file as a partial/in-progress export
    """
    if not products:
        return None
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Use exact template columns only
    all_keys = CSV_TEMPLATE_COLUMNS.copy()

    # Style definitions — clean, default Excel look
    header_fill = PatternFill("solid", fgColor="FFFFFF")
    header_font = Font(bold=True, color="000000", size=11, name="Calibri")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Side(style='thin', color='D0D0D0')
    cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # Write header row
    ws.row_dimensions[1].height = 35
    for ci, k in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=ci, value=k)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border

    # Data row styles — no alternating colors, just plain white
    norm_fill = PatternFill("solid", fgColor="FFFFFF")
    data_font = Font(size=10, name="Calibri")
    link_font = Font(size=10, name="Calibri", color="0563C1", underline="single")

    # Write data rows
    for ri, prod in enumerate(products, 2):
        ws.row_dimensions[ri].height = 20
        for ci, k in enumerate(all_keys, 1):
            val = prod.get(k, '')
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = norm_fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            # Style URL columns as links
            is_url = k in URL_COLUMNS and str(val).startswith('http')
            cell.font = link_font if is_url else data_font

    # Set column widths
    for ci, k in enumerate(all_keys, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COLUMN_WIDTHS.get(k, 22)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Create Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 38

    # Calculate statistics using correct field names
    summary_rows = [
        ("Scrape Summary", ""),
        ("Status", "PARTIAL — scraping still in progress" if partial else "COMPLETE"),
        ("Website", base_url),
        ("Keyword", keyword),
        ("Total Products", len(products)),
        ("Date & Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Fields Captured", len(all_keys)),
        ("", ""),
        ("With Sale Price", sum(1 for p in products if p.get('Sale Price'))),
        ("With Main Image", sum(1 for p in products if p.get('Main Image'))),
        ("With Brand", sum(1 for p in products if p.get('Brand'))),
        ("With Description", sum(1 for p in products if p.get('Detailed Description'))),
        ("With SKU", sum(1 for p in products if p.get('SKU'))),
    ]

    for ri, (label, value) in enumerate(summary_rows, 1):
        ws2.row_dimensions[ri].height = 22
        cell_a = ws2.cell(row=ri, column=1, value=label)
        cell_b = ws2.cell(row=ri, column=2, value=value)
        if ri == 1:
            cell_a.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
            cell_a.fill = PatternFill("solid", fgColor="1A1A2E")
        elif ri == 2 and partial:
            cell_a.font = Font(bold=True, size=11, name="Calibri")
            cell_b.font = Font(bold=True, size=11, name="Calibri", color="E67E22")
        else:
            cell_a.font = Font(bold=True, size=11, name="Calibri")
            cell_b.font = Font(size=11, name="Calibri")

    # Generate output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = re.sub(r'[^\w]', '_', keyword)[:20]
    partial_tag = "_PARTIAL" if partial else ""
    filepath = os.path.join(outputs_dir, f"scrape_{safe_keyword}_{timestamp}{partial_tag}.xlsx")
    wb.save(filepath)
    return filepath
